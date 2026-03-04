import streamlit as st
import pandas as pd
import datetime
from openpyxl import load_workbook
import easyocr
import numpy as np
from PIL import Image
import os
from fpdf import FPDF # Toto musí být zde a jen jednou


# Nastavení stránky
st.set_page_config(page_title="Hesti Zbuzany", layout="wide")

# --- INICIALIZACE AI (EasyOCR) ---
@st.cache_resource
def load_ocr():
    return easyocr.Reader(['en'], gpu=False)

reader = load_ocr()

# --- DATA PRO VÝBĚRY ---
DATA_VOZIDLA = {
    "MAN": ["TGX", "TGS", "TGM", "TGL"],
    "KRONE": ["Profi Liner", "Mega Liner", "Cool Liner", "Dry Liner", "Box Liner", "Box Carrier"],
    "D-TEC": [],
    "Langendorf": [],
    "Meiller": [],
    "Nooteboom": [],
    "O.ME.P.S": []
}

KAT_MAP = {
    "MAN nové": "MAN nové",
    "Návěsy nové": "Návěsy nové",
    "MAN TGE": "MAN TGE",
    "Ojetá vozidla": "Ojetá vozidla",
    "Noční stání": "Vlastní"
}

# --- FUNKCE: ZÁPIS PŘÍJEZDU ---
def zapsat_do_excelu(list_name, vyrobce, druh, vin, kode, poznamka):
    try:
        filename = "sluzba.xlsx"
        wb = load_workbook(filename)
        sheet = wb[list_name]
        
        last_row = sheet.max_row
        if last_row == 1 and sheet.cell(row=1, column=1).value is None:
            new_row = 1
            poradove_cislo = 1
        else:
            new_row = last_row + 1
            try:
                posledni_id = sheet.cell(row=last_row, column=1).value
                poradove_cislo = int(posledni_id) + 1
            except:
                poradove_cislo = new_row
        
        sheet.cell(row=new_row, column=1).value = poradove_cislo
        sheet.cell(row=new_row, column=2).value = vyrobce
        sheet.cell(row=new_row, column=3).value = druh if druh else ""
        sheet.cell(row=new_row, column=4).value = vin
        sheet.cell(row=new_row, column=5).value = kode
        sheet.cell(row=new_row, column=6).value = poznamka
        sheet.cell(row=new_row, column=7).value = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
        
        wb.save(filename)
        wb.close()
        return True
    except Exception as e:
        st.error(f"Chyba zápisu: {e}")
        return False

# --- FUNKCE: ZÁPIS ODJEZDU (PŘESUN DO LISTU VYDÁNO) ---
def zapsat_odjezd(list_z_kategorie, vin_kod, cil_odjezdu):
    try:
        filename = "sluzba.xlsx"
        wb = load_workbook(filename)
        
        if "Vydáno" not in wb.sheetnames:
            ws_vyd = wb.create_sheet("Vydáno")
            # Přidán sloupec "Cíl odjezdu"
            ws_vyd.append(["ID", "Výrobce", "Druh", "VIN/WERK", "KÓDE", "Poznámka", "Čas příjezdu", "Cíl odjezdu", "Čas odjezdu"])
        
        sheet_source = wb[list_z_kategorie]
        sheet_vydano = wb["Vydáno"]
        
        target_row_idx = None
        for row in range(1, sheet_source.max_row + 1):
            val = sheet_source.cell(row=row, column=4).value
            if val and str(val).strip().upper() == vin_kod.strip().upper():
                target_row_idx = row
                break
        
        if target_row_idx:
            last_row_vydano = sheet_vydano.max_row
            if last_row_vydano == 1:
                nove_id_vydano = 1
            else:
                try:
                    posledni_id = sheet_vydano.cell(row=last_row_vydano, column=1).value
                    nove_id_vydano = int(posledni_id) + 1
                except:
                    nove_id_vydano = last_row_vydano
            
            # Sestavení dat pro historii
            row_data = [nove_id_vydano] 
            for col in range(2, 8): # Výrobce až Čas příjezdu
                row_data.append(sheet_source.cell(row=target_row_idx, column=col).value)
            
            # Vložíme CÍL a pak ČAS ODJEZDU
            row_data.append(cil_odjezdu)
            row_data.append(datetime.datetime.now().strftime("%d.%m.%Y %H:%M"))
            
            sheet_vydano.append(row_data)
            sheet_source.delete_rows(target_row_idx)
            
            wb.save(filename)
            wb.close()
            return True
        else:
            wb.close()
            return False
    except Exception as e:
        st.error(f"Chyba při odjezdu: {e}")
        return False

# --- GLOBÁLNÍ DESIGN ---
st.markdown("""
    <style>
        .stButton > button {
            height: 3.5em;
            font-size: 18px !important;
            font-weight: bold !important;
            border-radius: 8px !important;
            width: 100%;
        }
        
        /* Barva pro tlačítko Příjezd */
        div[data-testid="stVerticalBlock"] > div:nth-child(2) button {
        }
.viewerBadge_link__1SuYq, .st-emotion-cache-15zrgzn e16nr0p34 {
    display: none !important;
}

/* Univerzální řešení pro všechny nadpisy */
h1 a, h2 a, h3 a, h4 a, h5 a, h6 a {
    display: none !important;
}
    </style>
""", unsafe_allow_html=True)

if 'stranka' not in st.session_state:
    st.session_state.stranka = "prehled"
if 'nacteny_vin' not in st.session_state:
    st.session_state.nacteny_vin = ""

# --- STRÁNKA: SEZNAMY ---
def stranka_seznam(nazev_kategorie, excel_list):
    st.header(f"📋 Seznam: {nazev_kategorie}")
    if st.button("⬅ ZPĚT"):
        st.session_state.stranka = "prehled"; st.rerun()
    try:
        data = pd.read_excel("sluzba.xlsx", sheet_name=excel_list, engine='openpyxl')
        st.dataframe(data, use_container_width=True, hide_index=True)
    except:
        st.info("Seznam je momentálně prázdný nebo list neexistuje.")

# --- STRÁNKA: HLAVNÍ REPORT ---
def stranka_report():
    # Vytvoříme dva sloupce: jeden široký pro nadpis a jeden užší pro logo
    col_nadpis, col_logo = st.columns([3, 1])
    
    with col_nadpis:
        st.title("🚛 HESTI ZBUZANY - PŘEHLED AREÁLU")
    
    with col_logo:
        # use_container_width zajistí, že se logo hezky vejde do sloupce
        st.image("HESTI GROUP.png", use_container_width=True)  
    
    prehled_data = {}
    celkem = 0
    try:
        wb = load_workbook("sluzba.xlsx", read_only=True, data_only=True)
        for nazev_app, nazev_excel in KAT_MAP.items():
            stav = 0
            if nazev_excel in wb.sheetnames:
                sh = wb[nazev_excel]
                for r in range(2, sh.max_row + 1):
                    if sh.cell(row=r, column=1).value is not None:
                        stav += 1
            prehled_data[nazev_app] = stav
            celkem += stav
        wb.close()

        # --- NOVÉ ZOBRAZENÍ (MÍSTO TABULKY) ---
        st.write("### 📊 AKTUÁLNÍ STAV VOZIDEL")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown(f"**MAN NOVÉ**")
            st.title(f"{prehled_data.get('MAN nové', 0)}")
            st.divider() # Čára pod 1. řádkem
            
            st.markdown(f"**OJETÁ VOZIDLA**")
            st.title(f"{prehled_data.get('Ojetá vozidla', 0)}")
            st.divider() # Čára pod 2. řádkem

        with col2:
            st.markdown(f"**NÁVĚSY NOVÉ**")
            st.title(f"{prehled_data.get('Návěsy nové', 0)}")
            st.divider() # Čára pod 1. řádkem
            
            st.markdown(f"**NOČNÍ STÁNÍ**")
            st.title(f"{prehled_data.get('Noční stání', 0)}")
            st.divider() # Čára pod 2. řádkem

        with col3:
            st.markdown(f"**MAN TGE**")
            st.title(f"{prehled_data.get('MAN TGE', 0)}")
            st.divider() # Čára pod 1. řádkem
            
            st.markdown(f"**CELKEM**")
            st.title(f"👉 {celkem}")
            st.divider() # Čára pod 2. řádkem

        

        # --- TLAČÍTKO PRO PDF EXPORT ---
        st.write("### 🖨️ Protokol pro předání směny")
        # Pro PDF si data musíme nechat v původním formátu seznamu
        pdf_list = [{"KATEGORIE VOZIDEL": k, "AKTUÁLNÍ STAV": v} for k, v in prehled_data.items()]
        pdf_list.append({"KATEGORIE VOZIDEL": "CELKEM", "AKTUÁLNÍ STAV": celkem})
        
        pdf_bytes = vygenerovat_pdf(pdf_list)
        if pdf_bytes:
            st.download_button(
                label="📄 STÁHNOUT PROTOKOL (PDF)",
                data=pdf_bytes,
                file_name=f"Stav_arealu_{datetime.datetime.now().strftime('%d_%m_%Y')}.pdf",
                mime="application/pdf",
                use_container_width=True
            )

    except Exception as e:
        st.error(f"Nepodařilo se načíst stavy: {e}")

    st.divider()
    # ... tady pokračují tvoje sloupce c1, c2 se Seznamy a Pohyby ...    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        st.subheader("🔎 Seznamy")
        if st.button("🚛 MAN NOVÉ", use_container_width=True): st.session_state.stranka = "man_list"; st.rerun()
        if st.button("🏗️ NÁVĚSY NOVÉ", use_container_width=True): st.session_state.stranka = "navesy_list"; st.rerun()
        if st.button("🚐 MAN TGE", use_container_width=True): st.session_state.stranka = "tge_list"; st.rerun()
        if st.button("🚗 OJETÁ VOZIDLA", use_container_width=True): st.session_state.stranka = "ojeta_list"; st.rerun()
        if st.button("📜 HISTORIE VÝDEJŮ (Vydáno)", use_container_width=True): st.session_state.stranka = "vydano_list"; st.rerun()
    with c2:
        st.subheader("⚙️ Pohyby")
        # Přidáme barvu pomocí Streamlit parametrů, pokud to verze dovoluje, 
        # nebo zůstaneme u CSS, ale opravíme selektor:
        st.button("🟢 ZAEVIDOVAT PŘÍJEZD", key="btn_prijezd", use_container_width=True)
        st.button("🔴 EVIDOVAT ODJEZD", key="btn_odjezd", use_container_width=True)

# --- STRÁNKA: PŘÍJEZD ---
def stranka_prijezd():
    st.header("🟢 Příjezd vozidla")
    foto = st.camera_input("📸 Vyfoťte štítek (VIN / WERK)")
    if foto:
        with st.spinner("🧠 AI analyzuje fotku..."):
            img = Image.open(foto)
            img_np = np.array(img)
            vysledek = reader.readtext(img_np)
            nalezené_texty = [t[1] for t in vysledek if len(t[1]) > 4]
            if nalezené_texty:
                st.session_state.nacteny_vin = max(nalezené_texty, key=len).upper().replace(" ", "")
                st.success(f"✅ AI přečetla: {st.session_state.nacteny_vin}")

    with st.container(border=True):
        vybrana_kat = st.selectbox("Druh vozidla pro zápis", list(KAT_MAP.keys()))
        col1, col2 = st.columns(2)
        with col1:
            vyrobce = st.selectbox("Výrobce:", list(DATA_VOZIDLA.keys()))
            druh = st.selectbox("Druh:", DATA_VOZIDLA[vyrobce]) if DATA_VOZIDLA[vyrobce] else ""
        with col2:
            vin = st.text_input("VIN / WERK:", value=st.session_state.nacteny_vin).upper()
            kode = st.text_input("KÓDE:")
        poznamka = st.text_area("Poznámka:")
        
        b1, b2 = st.columns(2)
        with b1:
            if st.button("💾 ULOŽIT PŘÍJEZD", use_container_width=True, type="primary"):
                if vin and zapsat_do_excelu(KAT_MAP[vybrana_kat], vyrobce, druh, vin, kode, poznamka):
                    st.success("Zapsáno!")
                    st.session_state.nacteny_vin = ""
        with b2:
            if st.button("⬅ ZPĚT", use_container_width=True):
                st.session_state.nacteny_vin = ""
                st.session_state.stranka = "prehled"; st.rerun()

# --- STRÁNKA: ODJEZD ---
def stranka_odjezd():
    st.header("🔴 Odjezd vozidla (Výdej)")
    with st.container(border=True):
        kat_odjezd = st.selectbox("Kategorie, ze které vozidlo odjíždí:", list(KAT_MAP.keys()))
        vin_odjezd = st.text_input("Zadejte VIN / WERK pro výdej:").upper()
        
        # NOVÉ POVINNÉ POLE
        cil_odjezdu = st.text_input("Kam vozidlo odjíždí? (např. Zákazník, Pobočka XY, Servis...)*")
        
        st.write("---")
        c1, c2 = st.columns(2)
        with c1:
            if st.button("❌ POTVRDIT VÝDEJ", use_container_width=True, type="primary"):
                if not vin_odjezd:
                    st.warning("Zadejte prosím VIN.")
                elif not cil_odjezdu:
                    st.error("Musíte vyplnit CÍL ODJEZDU!")
                else:
                    if zapsat_odjezd(KAT_MAP[kat_odjezd], vin_odjezd, cil_odjezdu):
                        st.success(f"✅ Hotovo! VIN {vin_odjezd} odjel směr: {cil_odjezdu}.")
                    else:
                        st.error(f"❌ VIN {vin_odjezd} v kategorii {kat_odjezd} nebyl nalezen.")
        with c2:
            if st.button("⬅ ZPĚT", use_container_width=True):
                st.session_state.stranka = "prehled"; st.rerun()

# --- FUNKCE PRO GENEROVÁNÍ PDF ---
def vygenerovat_pdf(prehled_data):
    try:
        pdf = FPDF()
        pdf.add_page()
        
        # --- VLOŽENÍ LOGA (Vpravo nahoře) ---
        logo_path = os.path.join(os.getcwd(), "HESTI GROUP.png")
        if os.path.exists(logo_path):
            # Logo zůstává vpravo, aby nezasahovalo do textu
            pdf.image(logo_path, x=150, y=8, w=45)
        
        # --- NAHRÁNÍ ČESKÉHO FONTU (Arial) ---
        font_path = os.path.join(os.getcwd(), "arial.ttf")
        if os.path.exists(font_path):
            pdf.add_font("ArialCesky", "", font_path)
            pdf.set_font("ArialCesky", "", 12)
            f_style = "ArialCesky"
        else:
            pdf.set_font("Helvetica", "", 12)
            f_style = "Helvetica"

        # --- TITULEK (Zarovnáno na střed) ---
        pdf.ln(15) # Mezera odshora
        pdf.set_font(f_style, "", 18)
        pdf.cell(190, 10, "PROTOKOL STAVU AREÁLU", ln=True, align="C")
        
        pdf.set_font(f_style, "", 13)
        pdf.cell(190, 8, "HESTI GROUP - Zbuzany", ln=True, align="C")
        
        # Datum a čas (Zarovnáno na střed)
        pdf.set_font(f_style, "", 10)
        datum_cas = datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        pdf.cell(190, 10, f"Vygenerováno: {datum_cas}", ln=True, align="C")
        
        pdf.ln(10) # Mezera před tabulkou
        
        # --- TABULKA (Zarovnáno na střed) ---
        # Aby byla tabulka na středu, musíme ji trochu odsadit zleva (190mm šířka listu - 170mm tabulka = 10mm okraj)
        pdf.set_x(20) 
        
        pdf.set_font(f_style, "", 11)
        pdf.set_fill_color(230, 230, 230) # Šedá hlavička
        pdf.cell(110, 10, "KATEGORIE VOZIDEL", border=1, fill=True, align="L")
        pdf.cell(60, 10, "STAV", border=1, ln=True, align="C", fill=True)
        
        # Data tabulky
        for radek in prehled_data:
            pdf.set_x(20) # Každý řádek musí začít na stejném odsazení
            
            # Zvýraznění řádku CELKEM
            is_total = radek["KATEGORIE VOZIDEL"] == "CELKEM"
            if is_total:
                pdf.set_fill_color(245, 245, 245)
            
            pdf.cell(110, 10, str(radek["KATEGORIE VOZIDEL"]), border=1, fill=is_total)
            pdf.cell(60, 10, str(radek["AKTUÁLNÍ STAV"]), border=1, ln=True, align="C", fill=is_total)

        # --- PODPISY (Zarovnané na střed) ---
        pdf.ln(40) # Mezera dolů k podpisům
        
        # Linky pro podpisy
        pdf.cell(95, 10, "__________________________", ln=0, align="C")
        pdf.cell(95, 10, "__________________________", ln=1, align="C")
        
        # Texty pod linkami
        pdf.set_font(f_style, "", 10)
        pdf.cell(95, 5, "Předal (podpis)", ln=0, align="C")
        pdf.cell(95, 5, "Převzal (podpis)", ln=1, align="C")
        
        return bytes(pdf.output())
        
    except Exception as e:
        st.error(f"Chyba při tvorbě PDF: {e}")
        return None


# --- ROZCESTNÍK ---
if st.session_state.stranka == "prehled":
    stranka_report()
elif st.session_state.stranka == "prijezd":
    stranka_prijezd()
elif st.session_state.stranka == "odjezd":
    stranka_odjezd()
else:
    mapping = {
        "man_list": ("MAN nové", "MAN nové"),
        "navesy_list": ("Návěsy nové", "Návěsy nové"),
        "tge_list": ("MAN TGE", "MAN TGE"),
        "ojeta_list": ("Ojetá vozidla", "Ojetá vozidla"),
        "vydano_list": ("Historie výdejů", "Vydáno")
    }
    info = mapping.get(st.session_state.stranka)
    if info:
        stranka_seznam(info[0], info[1])